"""Extract text from uploaded files. This is the CPU-bound work that runs in the
file worker (prefork pool).

Supported: PDF (with OCR fallback for scanned PDFs), DOCX, XLSX, CSV, images."""

import csv
import io

from app.core.logging import get_logger

log = get_logger("extractors")


def _ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def extract_text(filename: str, content_type: str, data: bytes) -> str:
    ext = _ext(filename)

    if ext == "pdf" or "pdf" in content_type:
        return _from_pdf(data)
    if ext in ("docx",) or "word" in content_type:
        return _from_docx(data)
    if ext in ("xlsx", "xlsm") or "spreadsheet" in content_type:
        return _from_xlsx(data)
    if ext == "csv" or "csv" in content_type:
        return _from_csv(data)
    if ext in ("png", "jpg", "jpeg", "tiff", "bmp", "webp") or content_type.startswith("image/"):
        return _from_image(data)

    # Fallback: try to decode as text.
    try:
        return data.decode("utf-8", errors="ignore")
    except Exception:
        return ""


def _from_pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)

    # If almost no text came out, it's probably a scanned PDF -> OCR it.
    if len(text.strip()) < 40:
        log.info("PDF looks scanned, falling back to OCR")
        return _ocr_pdf(data)
    return text


def _ocr_pdf(data: bytes) -> str:
    from pdf2image import convert_from_bytes
    import pytesseract

    pages = convert_from_bytes(data, dpi=200)
    return "\n".join(pytesseract.image_to_string(p) for p in pages)


def _from_docx(data: bytes) -> str:
    import docx

    document = docx.Document(io.BytesIO(data))
    return "\n".join(p.text for p in document.paragraphs)


def _from_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            lines.append("\t".join("" if c is None else str(c) for c in row))
    return "\n".join(lines)


def _from_csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="ignore")
    rows = list(csv.reader(io.StringIO(text)))
    return "\n".join("\t".join(r) for r in rows)


def _from_image(data: bytes) -> str:
    from PIL import Image
    import pytesseract

    img = Image.open(io.BytesIO(data))
    return pytesseract.image_to_string(img)
